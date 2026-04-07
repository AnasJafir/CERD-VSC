import os
import json
import re
import openpyxl

EXCEL_PATH = 'data/Code sources (1000 comptes).xlsx'
OUTPUT_JSON = 'config/sources_data.json'

def extract_aliases(nom):
    aliases = []
    nom = str(nom).strip()
    
    # Extraire ce qui est entre parenthèses (LVE) -> alias LVE
    parentheses_matches = re.findall(r'\(([^)]+)\)', nom)
    for match in parentheses_matches:
        alias = match.strip()
        if len(alias) >= 2 and len(alias) <= 15: # Ignorer les trop longues phrases
            aliases.append(alias)
            
    # Ajouter le nom complet sans les parenthèses
    clean_name = re.sub(r'\(.*?\)', '', nom).strip()
    if clean_name and clean_name != nom:
        aliases.append(clean_name)
        
    return aliases

def deduce_category(code_str):
    try:
        code = int(code_str)
        if 1 <= code <= 99: return "DIVERS"
        elif 100 <= code <= 199: return "Administration, Institutions, Associations (Maroc)"
        elif 200 <= code <= 299: return "Administration, Institutions, Associations (Étranger)"
        elif 300 <= code <= 649: return "Presse Maroc"
        elif 650 <= code <= 999: return "Presse Étranger"
    except ValueError:
        pass
    return "Inconnue"

def parse_excel():
    print(f"Ouverture du fichier {EXCEL_PATH} ...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    
    sources = []
    
    for row in range(1, ws.max_row + 1):
        code_cell = ws.cell(row=row, column=1).value
        nom_cell = ws.cell(row=row, column=2).value
        
        if code_cell is not None and nom_cell is not None:
            code_str = str(code_cell).strip()
            nom_str = str(nom_cell).strip()
            
            # Si le code est purement numérique
            if code_str.isdigit():
                obj = {
                    "code": code_str,
                    "nom": nom_str,
                    "categorie": deduce_category(code_str),
                    "aliases": extract_aliases(nom_str)
                }
                sources.append(obj)
                
    # Ajouter la source inconnue par défaut
    if not any(s['code'] == '999' for s in sources):
        sources.append({
            "code": "999",
            "nom": "Source inconnue",
            "categorie": "Presse Étranger",
            "aliases": ["Inconnue", "Non spécifié", "Source non listée"]
        })
        
    print(f"{len(sources)} sources extraites.")
    
    os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump({"sources": sources}, f, indent=2, ensure_ascii=False)
        
    print(f"✅ Fichier JSON généré avec succès dans : {OUTPUT_JSON}")
    
    print("\n--- INSTRUCTIONS POUR AIRTABLE ---")
    print("Veuillez créer une nouvelle table nommée 'Sources' avec les colonnes suivantes :")
    print(" 1. 'Code_Source' (Texte court) -> PRIMARY FIELD")
    print(" 2. 'Nom_Source' (Texte court)")
    print(" 3. 'Categorie' (Sélection unique / Texte)")
    print(" 4. 'Aliases' (Texte long)")
    
if __name__ == "__main__":
    parse_excel()
