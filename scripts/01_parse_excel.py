#!/usr/bin/env python3
"""
Script 01 — Parsing et normalisation (Exclusivité Stricte)
==========================================================
Régle d'Exclusivité :
  Un code est soit une Structure (Container), soit un Thème (Feuille).
  Jamais les deux.

Logique de triage :
  1. Domaines (X) -> Toujours Table 1.
  2. Secteurs (XX) :
     - Si enfants présents -> Table 2 (Structure).
     - Si zéro enfant    -> Table 3 (Thème).
  3. Sous-Secteurs (XX.X) :
     - Si enfants présents -> Table 2 (Structure).
     - Si zéro enfant    -> Table 3 (Thème).
  4. Thèmes (XX.XX) -> Toujours Table 3 (Thème).

Doublons : Aucun code ne doit apparaître dans plusieurs listes.
"""

import json
import openpyxl
from pathlib import Path

# ─── Chemins ──────────────────────────────────────────────────────────────────
ROOT_DIR = Path(__file__).resolve().parent.parent
EXCEL_PATH = ROOT_DIR / "data" / "Tables B3D_V1_Excel-3.3.26 Corrigé.xlsx"
OUTPUT_PATH = ROOT_DIR / "config" / "parsed_data.json"

CODE_FIXES = {
    "9.04": "90.4",
}

def normalize_code(raw_code) -> str | None:
    if raw_code is None: return None
    if isinstance(raw_code, float) and raw_code.is_integer():
         code_str = str(int(raw_code))
    else:
        code_str = str(raw_code).strip()
    
    if not code_str: return None
    # Suppression des en-têtes ou artefacts
    if code_str.lower() in ("code", "niveau", "theme", "secteur"):
        return None
        
    return CODE_FIXES.get(code_str, code_str)

def clean_text(value) -> str | None:
    if value is None: return None
    cleaned = str(value).strip()
    return cleaned if cleaned else None

def parse_excel(filepath: Path) -> dict:
    print(f"\nLecture Strict : {filepath.name}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Feuil1']

    # 1. Extraction à plat des noeuds
    nodes = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        code = normalize_code(row[0])
        if not code: continue
        
        # Heuristique nom : Col C prioritaire pour niveaux profonds
        nom = clean_text(row[2]) if '.' in code else clean_text(row[1])
        if not nom: nom = clean_text(row[1]) or clean_text(row[2]) or f"Code {code}"
        
        nodes[code] = {
            'code': code,
            'nom': nom,
            'children': set()
        }

    all_codes = sorted(nodes.keys())

    # 2. Construction de l'arbre (Parenté) -- QUI EST PARENT DE QUI ?
    for code in all_codes:
        parent_candidate = None
        
        # XX.Y ou XX.YY
        if '.' in code:
            parts = code.split('.')
            main, sub = parts[0], parts[1]
            
            # XX.X -> Parent XX
            if len(sub) == 1: 
                if main in nodes: parent_candidate = main
            
            # XX.XX -> Parent XX.X
            elif len(sub) >= 2: 
                potential_ss = f"{main}.{sub[0]}"
                if potential_ss in nodes:
                    parent_candidate = potential_ss
                elif main in nodes:
                    # Gestion Orphelin : XX.XX sans XX.X -> on attache à XX
                    # print(f"  [FIX] Orphelin {code} -> rattaché à {main}")
                    parent_candidate = main
        
        # XX -> Parent X (Domaine)
        elif len(code) == 2 and code.isdigit():
             dom_digit = code[0]
             if dom_digit in nodes:
                 parent_candidate = dom_digit

        if parent_candidate:
            nodes[parent_candidate]['children'].add(code)

    # 3. Distribution Exclusive
    final_data = {
        'domaines': [],
        'secteurs': [],
        'sous_secteurs': [],
        'themes': [],
        '_meta': {}
    }

    processed_codes = set()

    for code in all_codes:
        node = nodes[code]
        has_kids = len(node['children']) > 0
        
        # A. Domaine (Chiffre unique)
        if len(code) == 1 and code.isdigit():
            final_data['domaines'].append({'code': code, 'nom': node['nom']})
            processed_codes.add(code)
            continue
            
        # B. XX
        if '.' not in code and len(code) >= 2:
            domaine_bg = code[0]
            if has_kids:
                # Structure Secteur
                final_data['secteurs'].append({
                    'code': code, 
                    'nom': node['nom'], 
                    'parent_domaine': domaine_bg
                })
            else:
                # Terminal Thème
                final_data['themes'].append({
                    'code': code, 
                    'nom': node['nom'],
                    'type_parent': 'Domaine', 
                    'parent_ref': domaine_bg
                })
            processed_codes.add(code)
            continue

        # C. XX...
        if '.' in code:
            parts = code.split('.')
            main, sub = parts[0], parts[1]
            
            # XX.X
            if len(sub) == 1:
                parent_sect = main
                if has_kids:
                    # Structure Sous-Secteur
                    final_data['sous_secteurs'].append({
                        'code': code, 
                        'nom': node['nom'],
                        'parent_secteur': parent_sect,
                        'parent_domaine': parent_sect[0]
                    })
                else:
                    # Terminal Thème
                    final_data['themes'].append({
                        'code': code, 
                        'nom': node['nom'],
                        'type_parent': 'Secteur', 
                        'parent_ref': parent_sect
                    })
                processed_codes.add(code)
                continue
            
            # D. XX.XX (Toujours Thème)
            else: # len >= 2
                # On doit retrouver le parent effectif qu'on a utilisé
                potential_ss = f"{main}.{sub[0]}"
                
                real_parent = None
                p_type = None

                # Si XX.X existe et a des enfants (dont nous), c'est le parent
                if potential_ss in nodes and len(nodes[potential_ss]['children']) > 0:
                    real_parent = potential_ss
                    p_type = 'Sous_Secteur'
                elif main in nodes:
                    real_parent = main
                    p_type = 'Secteur'
                
                final_data['themes'].append({
                    'code': code, 
                    'nom': node['nom'],
                    'type_parent': p_type, # 'Sous_Secteur' ou 'Secteur'
                    'parent_ref': real_parent
                })
                processed_codes.add(code)

    # Stats
    total_parsed = len(processed_codes)
    print("\n  RÉPARTITION EXCLUSIVE :")
    print(f"  Domaines      : {len(final_data['domaines'])}")
    print(f"  Secteurs      : {len(final_data['secteurs'])}")
    print(f"  Sous-Secteurs : {len(final_data['sous_secteurs'])}")
    print(f"  Themes        : {len(final_data['themes'])}")
    print(f"  Total Unique  : {total_parsed} / {len(all_codes)}")
    
    final_data['_meta'] = {
        'counts': {
            'domaines': len(final_data['domaines']),
            'secteurs': len(final_data['secteurs']),
            'sous_secteurs': len(final_data['sous_secteurs']),
            'themes': len(final_data['themes'])
        }
    }

    if total_parsed != len(all_codes):
        print("  [ERREUR] Incohérence dans le comptage !")

    return final_data

if __name__ == '__main__':
    data = parse_excel(EXCEL_PATH)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"  Fichier écrit : {OUTPUT_PATH}")
